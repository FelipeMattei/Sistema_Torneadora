"""
Módulo Gerador de Excel (Relatórios Profissionais).

Gera relatórios em Excel (.xlsx) com:
- Cabeçalho estilizado com nome do sistema, título, período e saldo
- Tabela de dados formatada com filtros automáticos
- Freeze panes para manter cabeçalhos visíveis
- Colunas auto-dimensionadas
- Formatação de datas (dd/MM/yyyy) e moeda (R$)
"""

from typing import List, Tuple, Any, Optional
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def gerar_excel_relatorio(
    caminho_saida: str,
    titulo: str,
    periodo_descricao: str,
    saldo_final: str,
    colunas: List[str],
    linhas: List[Tuple],
    *,
    nome_sistema: str = "Sistema Financeiro - Torneadora",
    emitido_em: Optional[datetime] = None,
) -> bool:
    """
    Gera um arquivo Excel com layout profissional de relatório.
    
    Args:
        caminho_saida: Caminho completo onde o Excel será salvo (.xlsx)
        titulo: Título do relatório
        periodo_descricao: Descrição do período filtrado
        saldo_final: Texto do saldo/total final formatado
        colunas: Lista com nomes das colunas
        linhas: Lista de tuplas com os dados
        nome_sistema: Nome do sistema para o cabeçalho
        emitido_em: Data/hora de emissão (default: agora)
        
    Returns:
        True se gerado com sucesso, False caso contrário
    """
    try:
        emitido_em = emitido_em or datetime.now()
        total_registros = len(linhas)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # ===== ESTILOS =====
        
        # Fonte para título do sistema
        font_sistema = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        
        # Fonte para título do relatório
        font_titulo = Font(name='Calibri', size=14, bold=True, color='333333')
        
        # Fonte para metadados
        font_meta = Font(name='Calibri', size=11, color='555555')
        font_meta_bold = Font(name='Calibri', size=11, bold=True, color='333333')
        
        # Estilo do saldo (destaque)
        font_saldo = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        fill_saldo = PatternFill(start_color='00B33C', end_color='00B33C', fill_type='solid')
        
        # Estilo para cabeçalho da tabela
        font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        
        # Estilo para linhas alternadas
        fill_zebra = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Bordas
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # ===== CABEÇALHO DO RELATÓRIO =====
        
        row = 1
        
        # Nome do sistema
        ws.cell(row=row, column=1, value=nome_sistema)
        ws.cell(row=row, column=1).font = font_sistema
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
        row += 1
        
        # Título do relatório
        ws.cell(row=row, column=1, value=titulo)
        ws.cell(row=row, column=1).font = font_titulo
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colunas))
        row += 2
        
        # Período
        ws.cell(row=row, column=1, value="Período:")
        ws.cell(row=row, column=1).font = font_meta_bold
        ws.cell(row=row, column=2, value=periodo_descricao)
        ws.cell(row=row, column=2).font = font_meta
        row += 1
        
        # Data de emissão
        ws.cell(row=row, column=1, value="Emitido em:")
        ws.cell(row=row, column=1).font = font_meta_bold
        ws.cell(row=row, column=2, value=emitido_em.strftime("%d/%m/%Y às %H:%M"))
        ws.cell(row=row, column=2).font = font_meta
        row += 1
        
        # Total de registros
        ws.cell(row=row, column=1, value="Total de registros:")
        ws.cell(row=row, column=1).font = font_meta_bold
        ws.cell(row=row, column=2, value=total_registros)
        ws.cell(row=row, column=2).font = font_meta
        row += 2
        
        # Saldo final (destacado)
        saldo_txt = saldo_final.strip()
        ws.cell(row=row, column=1, value=saldo_txt)
        ws.cell(row=row, column=1).font = font_saldo
        ws.cell(row=row, column=1).fill = fill_saldo
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        # Merge para o saldo ficar mais visível
        merge_cols = min(3, len(colunas))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
        for col in range(1, merge_cols + 1):
            ws.cell(row=row, column=col).fill = fill_saldo
        row += 2
        
        # ===== TABELA DE DADOS =====
        
        header_row = row
        
        # Cabeçalho da tabela
        for col_idx, col_name in enumerate(colunas, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        row += 1
        
        # Dados
        for linha_idx, linha in enumerate(linhas):
            for col_idx, valor in enumerate(linha, start=1):
                cell = ws.cell(row=row, column=col_idx)
                
                # Formata o valor
                if isinstance(valor, float):
                    # Formato moeda brasileiro
                    cell.value = valor
                    cell.number_format = 'R$ #,##0.00'
                elif isinstance(valor, date) and not isinstance(valor, datetime):
                    cell.value = valor
                    cell.number_format = 'DD/MM/YYYY'
                elif isinstance(valor, datetime):
                    cell.value = valor
                    cell.number_format = 'DD/MM/YYYY HH:MM'
                elif valor is None:
                    cell.value = ""
                else:
                    cell.value = str(valor)
                
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
                
                # Zebra striping (linhas alternadas)
                if linha_idx % 2 == 1:
                    cell.fill = fill_zebra
            
            row += 1
        
        # ===== FILTROS AUTOMÁTICOS =====
        
        # Define a área da tabela para auto-filtro
        last_col_letter = get_column_letter(len(colunas))
        filter_range = f"A{header_row}:{last_col_letter}{row - 1}"
        ws.auto_filter.ref = filter_range
        
        # ===== FREEZE PANES =====
        
        # Congela linhas acima da tabela + cabeçalho da tabela
        ws.freeze_panes = f"A{header_row + 1}"
        
        # ===== AUTO-DIMENSIONAR COLUNAS =====
        
        for col_idx in range(1, len(colunas) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            # Verifica o cabeçalho
            header_value = colunas[col_idx - 1]
            max_length = max(max_length, len(str(header_value)) + 2)
            
            # Verifica os dados (até 100 linhas para performance)
            for linha_idx, linha in enumerate(linhas[:100]):
                if col_idx <= len(linha):
                    valor = linha[col_idx - 1]
                    if isinstance(valor, float):
                        cell_len = len(f"R$ {valor:,.2f}")
                    elif isinstance(valor, date):
                        cell_len = 12  # dd/mm/yyyy
                    elif valor is not None:
                        cell_len = len(str(valor))
                    else:
                        cell_len = 0
                    max_length = max(max_length, cell_len)
            
            # Aplica largura (com margem)
            adjusted_width = min(max_length + 3, 50)  # Max 50 caracteres
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # ===== SALVAR =====
        
        wb.save(caminho_saida)
        return True
        
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        return False
